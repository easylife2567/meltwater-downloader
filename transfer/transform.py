"""将 Meltwater 原始下载数据转换为目标 Excel 格式。

用法:
    python -m transfer.transform <输入文件> [--output <输出文件>] [--name <Input Name>]

示例:
    python -m transfer.transform data/meltwater_feed_20260728_001835.xlsx
    python -m transfer.transform data/meltwater_feed_20260728_001835.xlsx --name "我的搜索"
    python -m transfer.transform data/meltwater_feed_20260728_001835.xlsx -o output/result.xlsx
"""

import argparse
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

import openpyxl
from openpyxl.utils import get_column_letter

# ── 字段映射定义 ──
# 目标列名 → (Meltwater 源字段, 默认值, 转换函数)
# 转换函数签名为: (value, row_dict) -> Any

_COUNTRY_MAP = {
    "cn": "China",
    "us": "United States",
    "gb": "United Kingdom",
    "jp": "Japan",
    "kr": "South Korea",
    "de": "Germany",
    "fr": "France",
    "tw": "Taiwan",
    "hk": "Hong Kong",
}


def _fmt_date(v: Any, _row: dict) -> Optional[str]:
    """ISO 8601 → YYYY-MM-DD"""
    if not v:
        return None
    try:
        dt = datetime.fromisoformat(str(v).replace("Z", "+00:00"))
        return dt.strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return str(v)[:10]


def _fmt_alt_date(v: Any, _row: dict) -> Optional[str]:
    """ISO 8601 → DD/MM/YYYY HH:MM"""
    if not v:
        return None
    try:
        dt = datetime.fromisoformat(str(v).replace("Z", "+00:00"))
        return dt.strftime("%d/%m/%Y %H:%M")
    except (ValueError, TypeError):
        return str(v)[:16]


def _fmt_time(v: Any, _row: dict) -> Optional[str]:
    """ISO 8601 → HH:MM:SS"""
    if not v:
        return None
    try:
        dt = datetime.fromisoformat(str(v).replace("Z", "+00:00"))
        return dt.strftime("%H:%M:%S")
    except (ValueError, TypeError):
        return None


def _country(v: Any, _row: dict) -> Optional[str]:
    """国家代码 → 英文名"""
    if not v:
        return None
    return _COUNTRY_MAP.get(str(v).lower(), str(v))


def _influencer(v: Any, row: dict) -> Optional[str]:
    """提取有意义的作者名：优先 author_handle，gh_ 开头的用 author_name"""
    handle = row.get("author_handle", "")
    name = row.get("author_name", "")
    if handle and not str(handle).startswith("gh_"):
        return str(handle)
    if name and not str(name).startswith("gh_"):
        return str(name)
    return str(handle or name or "")


def _open_text(v: Any, row: dict) -> Optional[str]:
    """Opening Text：取 content 前 300 字"""
    text = v or row.get("full_content") or row.get("ingress") or ""
    return str(text)[:300] if text else None


def _sum_fields(*fields: str):
    """生成一个求和函数，对多个字段求和"""
    def inner(_v: Any, row: dict) -> int:
        total = 0
        for f in fields:
            val = row.get(f, 0) or 0
            try:
                total += int(val)
            except (ValueError, TypeError):
                pass
        return total
    return inner


def _engagement(_v: Any, row: dict) -> int:
    """综合互动量：所有社交指标之和"""
    metrics = [
        "tw_retweets", "tw_likes", "tw_replies",
        "fb_likes", "fb_shares", "fb_post_reactions",
    ]
    total = 0
    for f in metrics:
        val = row.get(f, 0) or 0
        try:
            total += int(val)
        except (ValueError, TypeError):
            pass
    return total


# ── 映射表 ──
# 每个元素: (目标列名, 源字段 or None, 转换函数 or 静态值)
# 若源字段为 None 且转换函数为可调用 → 转换函数从整行取值
# 若转换函数不是可调用 → 当作静态默认值
FIELD_MAP = [
    # (目标列, 源字段/None, 转换函数)
    ("Date",              "published_at",     _fmt_date),
    ("Headline",          "title",            None),
    ("URL",               "url",              None),
    ("Opening Text",      "content",          _open_text),
    ("Hit Sentence",      "match_sentence",   None),
    ("Source",            "source",           None),
    ("Influencer",        None,               _influencer),
    ("Country",           "country",          _country),
    ("Subregion",         "region",           None),
    ("Language",          "language",         None),
    ("Reach",             "reach",            None),
    ("Desktop Reach",     "local_reach",      None),
    ("Mobile Reach",      "global_reach",     None),
    ("Twitter Social Echo", None,             _sum_fields("tw_retweets", "tw_likes", "tw_replies")),
    ("Facebook Social Echo", None,            _sum_fields("fb_likes", "fb_shares", "fb_post_reactions")),
    ("Reddit Social Echo", None,              0),
    ("National Viewership", None,              None),
    ("Engagement",        None,               _engagement),
    ("AVE",               "ave",              None),
    ("Sentiment",         "sentiment",        None),
    ("Key Phrases",       "key_phrases",      None),
    ("Input Name",        None,               None),  # 由命令行 --name 指定
    ("Keywords",          "keywords",         None),
    ("Twitter Authority", "author_authority",  None),
    ("Tweet Id",          "external_id",      None),
    ("Twitter Id",        None,               None),
    ("Twitter Client",    None,               None),
    ("Twitter Screen Name", "author_handle",  None),
    ("User Profile Url",  "author_url",       None),
    ("Twitter Bio",       None,               None),
    ("Twitter Followers", "tw_followers",     None),
    ("Twitter Following", "tw_following",     None),
    ("Alternate Date Format", "published_at", _fmt_alt_date),
    ("Time",              "published_at",     _fmt_time),
    ("State",             None,               None),
    ("City",              "place",            None),
    ("Social Echo Total", None,               None),  # 需要计算时可在后处理
    ("Editorial Echo",    None,               None),
    ("Views",             None,               None),
    ("Estimated Views",   None,               None),
    ("Likes",             "tw_likes",         None),
    ("Replies",           "tw_replies",       None),
    ("Retweets",          "tw_retweets",      None),
    ("Comments",          None,               None),
    ("Shares",            "fb_shares",        None),
    ("Reactions",         "fb_post_reactions", None),
    ("Threads",           None,               None),
    ("Is Verified",       "author_verified",  None),
    ("Parent URL",        "source_url",       None),
    ("Document Tags",     None,               None),
    ("Document ID",       "document_id",      None),
]


def transform_row(src_row: dict, input_name: str = "") -> dict:
    """将一行 Meltwater 数据转换为目标格式。

    Args:
        src_row: Meltwater 原始行（列名 → 值）
        input_name: 填入 Input Name 列的值

    Returns:
        目标格式的行（列名 → 值）
    """
    result = {}
    for target_col, src_field, xform in FIELD_MAP:
        if target_col == "Input Name":
            result[target_col] = input_name
            continue

        if src_field is not None:
            raw = src_row.get(src_field, "")

            if callable(xform):
                result[target_col] = xform(raw, src_row)
            else:
                result[target_col] = raw
        else:
            # 无源字段 — xform 可能是函数（从整行计算）或静态值
            if callable(xform):
                result[target_col] = xform(None, src_row)
            else:
                result[target_col] = xform

    # 计算 Social Echo Total
    te = result.get("Twitter Social Echo") or 0
    fe = result.get("Facebook Social Echo") or 0
    re = result.get("Reddit Social Echo") or 0
    result["Social Echo Total"] = (int(te) + int(fe) + int(re)) if any([te, fe, re]) else None

    return result


def transform_file(
    input_path: str,
    output_path: Optional[str] = None,
    input_name: str = "",
) -> Path:
    """转换整个 xlsx 文件。

    Args:
        input_path: 源 Meltwater xlsx 文件路径
        output_path: 输出路径；为 None 时自动生成
        input_name: Input Name 列的值

    Returns:
        输出文件路径
    """
    src = Path(input_path)
    if not src.exists():
        raise FileNotFoundError(f"输入文件不存在: {input_path}")

    # 读取源文件
    wb_src = openpyxl.load_workbook(src)
    ws_src = wb_src.active
    src_headers = [cell.value for cell in ws_src[1]]

    # 构建目标表头
    target_headers = [col for col, _, _ in FIELD_MAP]

    # 创建目标工作簿
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "data"

    # 写入表头
    for ci, h in enumerate(target_headers, 1):
        cell = ws_out.cell(row=1, column=ci, value=h)
        cell.font = openpyxl.styles.Font(bold=True)

    # 写入数据行
    for ri in range(2, ws_src.max_row + 1):
        src_row = {}
        for ci, h in enumerate(src_headers):
            src_row[h] = ws_src.cell(row=ri, column=ci + 1).value

        target_row = transform_row(src_row, input_name)

        for ci, h in enumerate(target_headers, 1):
            ws_out.cell(row=ri, column=ci, value=target_row.get(h))

    # 自动列宽（取前 100 行估算）
    _auto_column_width(ws_out, target_headers)

    # 确定输出路径
    if output_path:
        dst = Path(output_path)
    else:
        dst = src.parent / f"{src.stem}_transformed.xlsx"
    dst.parent.mkdir(parents=True, exist_ok=True)

    wb_out.save(dst)
    print(f"✓ 转换完成: {ws_src.max_row - 1} 行 → {dst}")
    return dst


def _auto_column_width(ws, headers: list):
    """根据内容自动调整列宽（上限 60，下限 8）。"""
    sample_rows = min(ws.max_row, 100)
    for ci, h in enumerate(headers, 1):
        max_len = len(str(h))
        for ri in range(2, sample_rows + 1):
            val = ws.cell(row=ri, column=ci).value
            if val:
                max_len = max(max_len, min(len(str(val)), 40))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 60)


def main():
    parser = argparse.ArgumentParser(
        description="将 Meltwater 下载 xlsx 转换为目标格式",
    )
    parser.add_argument("input", help="Meltwater 源 xlsx 文件路径")
    parser.add_argument(
        "-o", "--output",
        help="输出文件路径（默认：同目录下 <原名>_transformed.xlsx）",
    )
    parser.add_argument(
        "--name",
        default="",
        help="Input Name 列的值（默认空）",
    )
    args = parser.parse_args()

    try:
        transform_file(args.input, args.output, args.name)
    except Exception as e:
        print(f"[ERROR] {e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
